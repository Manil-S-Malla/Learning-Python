import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def correctWorkbook(wb, sh):
    workbook = xl.load_workbook(wb)
    sheet = workbook[sh]

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * .9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
        print(corrected_price_cell.value)
    
    corrected_values = Reference(sheet, min_row= 2, max_row= sheet.max_row, min_col= 4, max_col= 4)
    chart= BarChart()
    chart.add_data(corrected_values)
    sheet.add_chart(chart, 'e2')

    workbook.save('corrected_transactions.xlsx')

correctWorkbook('transactions.xlsx', 'Sheet1')