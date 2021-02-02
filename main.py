from bs4 import BeautifulSoup
import requests
import time
from os import path
import openpyxl as xl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

URL = 'https://www.timesjobs.com/candidate/job-search.html?searchType=personalizedSearch&from=submit&txtKeywords=python&txtLocation='
date_filter = 'few days'

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
   
def newWorkbook(file_name):
    wb = xl.Workbook()
    sheet = wb.active

    sheet['A1'] = 'Company Name' 
    sheet['B1'] = 'Additional Information URL'
    sheet['C1'] = 'Skill Required'
    sheet['D1'] = 'Skill Required'
    sheet['E1'] = 'Skill Required'
    sheet['F1'] = 'Skill Required'
    sheet['G1'] = 'Skill Required'
    sheet['H1'] = 'Skill Required'
    sheet['I1'] = 'Skill Required'
    sheet['J1'] = 'Skill Required'

    wb.save(f"{file_name.split('.',1)[0]}.{file_name.split('.',1)[1]}")


def fileExists(file_name):
    if(path.exists(file_name)):
        return True
    else:
        return False
        

def getMultipleInputs(sentinel= '`'):
    for input_values in iter(input, sentinel):
        yield input_values


def getResponse(URL):
    html_response_text = requests.get(URL).text
    soup = BeautifulSoup(html_response_text, 'lxml')
    return soup.find_all('li', class_ = 'clearfix job-bx wht-shd-bx')    

def getUserSkills():
    print('\nInput your skills. [Press ` to exit.]')
    return list(getMultipleInputs())


def filterJobs(jobs, date_filter, user_skills):
    filtered_information = list()

    for job in jobs:
        company_name = job.find('h3', class_ = 'joblist-comp-name').text.replace('  ', '')
        skills_needed = job.find('span', class_ = 'srp-skills').text.replace('  ', '')
        posted_date = job.find('span', class_ = 'sim-posted').span.text.partition(' ')[2].partition(' ago')[0]
        additional_info_url = job.header.h2.a['href']

        if(posted_date == date_filter):
            for skill in user_skills:
                if(skill.lower() in skills_needed.lower()):
                    info = {
                        'Company Name' : company_name.strip(),
                        'Additional Info URL' : additional_info_url,
                        'Skills Required': skills_needed.strip().split(',')
                    }
                    filtered_information.append(info)
        
    return filtered_information
                    

def outputCsv(info, file_name = 'Job Posts.xlsx'):
    if not fileExists(file_name):
        newWorkbook(file_name)

    wb = xl.load_workbook(file_name)
    sheet = wb[wb.sheetnames[0]]
    
    last_cell = str(list(sheet.columns)[-1]).split('.')[1].split('>')[0]
    last_row = int(coordinate_from_string(last_cell)[1])

    for element in info:
        last_row += 1
        sheet.cell(row = last_row, column = 1, value = element['Company Name'])
        sheet.cell(row = last_row, column = 2, value = element['Additional Info URL']) 
        skill_col = 3
        for skill in element['Skills Required']:
            sheet.cell(row = last_row, column = skill_col, value = skill) 
            skill_col += 1  

    wb.save(file_name)
    print(f'\nFiltered information stored to {file_name}.')

def findJobs():
    print(f'''\n\t\t\t\t{bcolors.BOLD}{bcolors.OKGREEN}Times Jobs Scraper{bcolors.ENDC}{bcolors.ENDC}   
    \n   This is a web scraper which scrapes data from 'timesjobs.com' and filters the data based on posted date and skills.
    \n   Currently the following URL is selected.
    {bcolors.OKCYAN}{URL}{bcolors.ENDC} ''')
    
    changeUrl = input('\nChange URL ? [y/n] ')
    if (changeUrl == 'y'):
        url = input('\nInput new URL: ')
    else:
        url = URL

    filtered_jobs = filterJobs(getResponse(url), date_filter, getUserSkills())
    outputCsv(filtered_jobs, 'Job Posts.xlsx')


if __name__ == '__main__':
    while True:
        findJobs()

        waiting_time_min = int(input('\nRepeat every ? mins. [0 or less mins for no repeat] '))
        if (waiting_time_min > 0):
            print(f'\nRunning every {waiting_time_min} mins.')
            time.sleep(waiting_time_min * 60)
        else:
            break